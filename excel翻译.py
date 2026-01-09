"""
Excel文档智能翻译工具
功能：自动翻译中文Excel文档为英文，将结果换行追加到原文本单元格内，支持缓存、重试、多线程等特性
"""

# 标准库导入
import hashlib    # 用于生成MD5签名
import random     # 生成随机数
import requests   # 发送HTTP请求
import os         # 文件路径操作
import json       # 缓存文件读写
import argparse   # 命令行参数解析
import time       # 时间控制
import re         # 正则表达式处理
import threading  # 线程锁
from concurrent.futures import ThreadPoolExecutor, as_completed  #线程池管理
from openpyxl import load_workbook  # Excel文件操作
from openpyxl.styles import Font  # 字体设置
from tqdm import tqdm  # 进度条显示

# ========== 全局配置 ==========
appid = '20241223002235907'      # 百度翻译API应用ID（需自行申请）
secretKey = 'FaoqgT8hOE5JLFoySt9S'  # 百度翻译API密钥
url = 'https://fanyi-api.baidu.com/api/trans/vip/translate'  # API端点
MAX_RETRIES = 10    # 最大重试次数
MAX_WORKERS = 4     # 最大线程数
CACHE_FILE = 'translation_cache.json'  # 缓存文件路径

# ========== 初始化缓存与锁 ==========
# 加载翻译缓存（如果存在）
if os.path.exists(CACHE_FILE):
    with open(CACHE_FILE, 'r', encoding='utf-8') as f:
        translation_cache = json.load(f)
else:
    translation_cache = {}

# 创建线程锁
cache_lock = threading.Lock()   # 缓存访问锁
print_lock = threading.Lock()   # 控制台输出锁

def save_cache():
    """将缓存保存到文件（线程安全）"""
    with cache_lock:
        with open(CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(translation_cache, f, ensure_ascii=False, indent=2)

def baidu_translate(text, attempt):
    """
    调用百度翻译API进行单次翻译尝试
    :param text: 待翻译文本
    :param attempt: 当前尝试次数（用于错误提示）
    :return: 翻译结果或None
    """
    # 生成随机盐值
    salt = random.randint(32768, 65536)
    # 构造签名
    sign_str = appid + text + str(salt) + secretKey
    sign = hashlib.md5(sign_str.encode()).hexdigest()
    
    # 请求参数
    params = {
        'q': text,
        'from': 'zh',
        'to': 'en',
        'appid': appid,
        'salt': salt,
        'sign': sign
    }
    
    try:
        # 发送请求
        r = requests.get(url, params=params, timeout=10)
        r.raise_for_status()  # 检查HTTP状态码
        
        result = r.json()
        # 验证响应结构
        if 'trans_result' not in result:
            raise ValueError(f"Invalid response: {result}")
            
        return result['trans_result'][0]['dst']
    except Exception as e:
        # 输出错误信息（线程安全）
        with print_lock:
            tqdm.write(f"⚠️ 尝试 {attempt} 失败: {text[:20]}... ({str(e)})")
        return None

def baidu_translate_cached(text):
    """
    带缓存和重试机制的翻译函数
    :param text: 待翻译文本
    :return: 翻译结果或None
    """
    # 检查缓存（如果未禁用缓存）
    if not args.nocache:
        with cache_lock:
            if text in translation_cache:
                return translation_cache[text]

    # 特殊处理：直接映射"回归测试"
    if '回归测试' in text:
        translation_cache[text] = 'Regression testing'
        return 'Regression testing'

    # 常规翻译流程
    translated = None
    for attempt in range(1, MAX_RETRIES + 1):
        translated = baidu_translate(text, attempt)
        if translated:
            # 重试成功提示
            if attempt > 1:
                with print_lock:
                    tqdm.write(f"✅ 重试后成功: {text[:30]} (第 {attempt} 次)")
            # 更新缓存
            with cache_lock:
                translation_cache[text] = translated
            return translated
        # 指数退避策略
        time.sleep(2 ** min(attempt, 5) + random.random())

    return None

def split_into_sentences(text):
    """
    将文本分割成句子
    :param text: 原始文本
    :return: 句子列表
    """
    # 使用正则表达式分割句子：
    # - 在句号、感叹号、问号、分号后分割
    # - 在换行符前分割
    # - 在数字编号（如"1."）前分割
    return [s.strip() for s in re.split(r'(?<=[。！？；])|(?=\n)|(?=\d+\.)', text) if s.strip()]

def process_cell(cell, pbar, stats):
    """
    处理单个单元格（线程执行单元）
    :param cell: Excel单元格对象
    :param pbar: 进度条对象
    :param stats: 统计信息字典
    """
    text = str(cell.value).strip() if cell.value else ""
    if not text:
        pbar.update(1)
        return

    translated = baidu_translate_cached(text)
    if translated:
        # 将翻译结果换行追加到原文本单元格
        original_value = cell.value if cell.value else ""
        cell.value = f"{original_value}\n{translated}"
        stats['success'] += 1
    else:
        stats['failed'] += 1

    save_cache()
    pbar.update(1)

def translate_excel(file_path):
    """执行Excel翻译流程"""
    global translation_cache

    # 自动识别文件类型
    if not file_path.lower().endswith(('.xlsx', '.xls')):
        print("❌ 不支持的文件格式")
        return

    # 处理缓存相关参数
    if args.resetcache and os.path.exists(CACHE_FILE):
        os.remove(CACHE_FILE)
        translation_cache = {}
        print("🗑 已清空缓存文件")
    elif not os.path.exists(CACHE_FILE):
        translation_cache = {}
    elif not args.nocache:
        with open(CACHE_FILE, 'r', encoding='utf-8') as f:
            translation_cache = json.load(f)

    # 加载Excel文件
    print(f"📄 正在处理文档：{file_path}")
    wb = load_workbook(file_path)
    ws = wb.active

    # 收集所有需要翻译的单元格
    all_cells = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column, values_only=False):
        for cell in row:
            all_cells.append(cell)

    print(f"📝 发现可翻译单元格: {len(all_cells)} 个")
    
    # 执行翻译
    stats = {'success': 0, 'failed': 0}
    with tqdm(total=len(all_cells), desc="翻译进度", ncols=80) as pbar:
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = []
            # 提交翻译任务
            for cell in all_cells:
                future = executor.submit(process_cell, cell, pbar, stats)
                futures.append(future)

            # 等待任务完成
            for future in as_completed(futures):
                try:
                    future.result()
                except Exception as e:
                    with print_lock:
                        tqdm.write(f"⚠️ 线程错误: {str(e)}")

    # 保存结果文件
    base_name, ext = os.path.splitext(file_path)
    new_name = base_name + "_translate" + ext
    wb.save(new_name)
    print(f"\n✅ 翻译完成 - 成功: {stats['success']}, 失败: {stats['failed']} 条目")
    print(f"📁 文件已保存为：{new_name}")

if __name__ == "__main__":
    # 命令行参数解析
    parser = argparse.ArgumentParser(description="Excel文档智能翻译工具")
    parser.add_argument('-f', '--file', required=True, help='Excel文件路径')
    parser.add_argument('--nocache', action='store_true', help='忽略缓存并强制重新翻译')
    parser.add_argument('--resetcache', action='store_true', help='清空缓存后再翻译')
    args = parser.parse_args()
    
    # 启动翻译流程
    translate_excel(args.file)